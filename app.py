# app.py
import os
from flask import Flask, request, jsonify
from dotenv import load_dotenv
import openai
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path
import json

load_dotenv()

# OpenAI
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")  # change if you want gpt-4 or gpt-5
if OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY

# Email & Excel settings
EMAIL_SMTP_HOST = os.getenv("EMAIL_SMTP_HOST", "smtp.gmail.com")
EMAIL_SMTP_PORT = int(os.getenv("EMAIL_SMTP_PORT", "587"))
EMAIL_USERNAME = os.getenv("EMAIL_USERNAME", "")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")
LEADS_EMAIL_TO = os.getenv("LEADS_EMAIL_TO", "contact@besolarsolutions.com")
EXCEL_PATH = os.getenv("LEADS_EXCEL_PATH", "data/leads.xlsx")

app = Flask(__name__)

SYSTEM_PROMPT = os.getenv("SYSTEM_PROMPT", 
    "You are Be Solar Assistant. Friendly, expert. Identify intent, ask for city, residential/commercial, monthly bill or kW, then request contact info. Keep answers concise."
)

def ensure_excel(path):
    path_obj = Path(path)
    if not path_obj.parent.exists():
        path_obj.parent.mkdir(parents=True, exist_ok=True)
    if not path_obj.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        headers = ["received_at", "name", "phone", "email", "city", "type", "monthly_bill", "estimated_kw", "notes"]
        ws.append(headers)
        wb.save(path)

def append_to_excel(path, lead: dict):
    ensure_excel(path)
    wb = load_workbook(path)
    ws = wb.active
    row = [
        lead.get("received_at", ""),
        lead.get("name", ""),
        lead.get("phone", ""),
        lead.get("email", ""),
        lead.get("city", ""),
        lead.get("type", ""),
        lead.get("monthly_bill", ""),
        lead.get("estimated_kw", ""),
        lead.get("notes", "")
    ]
    ws.append(row)
    wb.save(path)

def send_email(subject, body, to_addr):
    if not EMAIL_USERNAME or not EMAIL_PASSWORD:
        app.logger.warning("Email credentials not configured. Skipping email send.")
        return False
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_USERNAME
        msg["To"] = to_addr
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))
        server = smtplib.SMTP(EMAIL_SMTP_HOST, EMAIL_SMTP_PORT, timeout=15)
        server.starttls()
        server.login(EMAIL_USERNAME, EMAIL_PASSWORD)
        server.sendmail(EMAIL_USERNAME, [to_addr], msg.as_string())
        server.quit()
        return True
    except Exception as e:
        app.logger.error(f"Email send failed: {e}")
        return False

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status":"ok"})

@app.route("/chat", methods=["POST"])
def chat():
    """
    Request JSON:
    { "message": "I need a 10kW system in Lahore" }
    """
    data = request.get_json(force=True)
    if not data or "message" not in data:
        return jsonify({"error":"no message provided"}), 400
    user_message = data["message"]

    if not OPENAI_API_KEY:
        return jsonify({"error":"OpenAI key not configured on server"}), 500

    messages = [
        {"role":"system", "content": SYSTEM_PROMPT},
        {"role":"user", "content": user_message}
    ]
    try:
        resp = openai.ChatCompletion.create(
            model=OPENAI_MODEL,
            messages=messages,
            max_tokens=400,
            temperature=0.2
        )
        reply = resp.choices[0].message.content
        return jsonify({"reply": reply})
    except Exception as e:
        app.logger.error("OpenAI error: %s", e)
        return jsonify({"error": str(e)}), 500

@app.route("/lead", methods=["POST"])
def lead():
    """
    Accepts JSON:
    {
      "name": "...",
      "phone":"....",
      "email":"...",
      "city":"Lahore",
      "type":"Residential",
      "monthly_bill":"20000",
      "estimated_kw":"10",
      "notes":"..."
    }
    """
    data = request.get_json(force=True)
    if not data:
        return jsonify({"error":"no JSON body"}), 400

    lead = {
        "received_at": datetime.utcnow().isoformat(),
        "name": data.get("name", ""),
        "phone": data.get("phone", ""),
        "email": data.get("email", ""),
        "city": data.get("city", ""),
        "type": data.get("type", ""),
        "monthly_bill": data.get("monthly_bill", ""),
        "estimated_kw": data.get("estimated_kw", ""),
        "notes": data.get("notes", "")
    }

    # Append to Excel (best-effort)
    try:
        append_to_excel(EXCEL_PATH, lead)
    except Exception as e:
        app.logger.error(f"Excel write failed: {e}")

    # Send email notification (best-effort)
    subject = f"New Solar Lead: {lead.get('name','-')} ({lead.get('city','-')})"
    body = json.dumps(lead, indent=2)
    email_ok = send_email(subject, body, LEADS_EMAIL_TO)

    return jsonify({"status":"ok", "email_sent": email_ok, "lead": lead})

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port)
